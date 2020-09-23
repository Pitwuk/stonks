import requests
from bs4 import BeautifulSoup
import openpyxl as xl
import time

from Scraper import Scraper

# This program saves the date, time, and text contents of all news articles with the search result for
# nvidia on tech radar into an xlsx file

start_time = time.time()
schtuff = []

# 137 pages (10 a peice)
iteration = 0
pages_per = 137
file = 'NVDA\\thevergeNVDAmasterhistory.xlsx'
worksheet = xl.load_workbook(file).active
newWB = xl.Workbook()
newSheet = newWB.active

for row in worksheet.iter_rows():
    row_arr = []
    for i in range(4):
        row_arr.append(row[i].value)
    newSheet.append(row_arr)

for i in range(pages_per):
    page = str(137-(i+(pages_per*iteration)))
    print('Page: ' + page)
    art1 = Scraper(
        'https://www.theverge.com/search?order=date&page=' + page + '&q=nvidia&type=Article', 'vg', 1)
    num_articles = art1.getNumArticles()
    if num_articles == 0:
        time.sleep(60)
        num_articles = art1.getNumArticles()
    for j in range(num_articles):
        art1 = Scraper('https://www.theverge.com/search?order=date&page=' +
                       page + '&q=nvidia&type=Article', 'vg', int(num_articles-j))
        try:
            schtuff.append([art1.getArtDate(), art1.getArtTime(),
                            art1.getArtLink(), art1.getArtFilteredText()])
        except:
            time.sleep(60)
            schtuff.append([art1.getArtDate(), art1.getArtTime(),
                            art1.getArtLink(), art1.getArtFilteredText()])
        print('article ' + str(j+((i+(pages_per*iteration))*10)) + '/' + str(1370))
        print(len(schtuff[j][3]))
for i in range(len(schtuff)):
    newSheet.append(schtuff[i])

newWB.save(file)

print("Finished iteration: " + str(iteration))
print('Elapsed Time: ' + str(time.time()-start_time))