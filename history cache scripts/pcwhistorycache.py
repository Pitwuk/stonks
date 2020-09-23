import requests
from bs4 import BeautifulSoup
import openpyxl as xl
import time

from Scraper import Scraper

# This program saves the date, time, and text contents of all news articles with the search result for
# nvidia on tech radar into an xlsx file

start_time = time.time()
schtuff = []

# 0-2150
iteration = 5
pages_per = 43
file = 'pcworldNVDAmasterhistory.xlsx'
worksheet = xl.load_workbook(file).active
newWB = xl.Workbook()
newSheet = newWB.active

for row in worksheet.iter_rows():
    row_arr = []
    for i in range(4):
        row_arr.append(row[i].value)
    newSheet.append(row_arr)

for i in range(pages_per):
    art1 = Scraper('https://www.pcworld.com/search?query=nvidia&s=d&start=' +
                   str(2150-10*(i+(pages_per*iteration))) + '&submit=search&t=2', 'pcw', 0)
    num_articles = art1.getNumArticles()
    for j in range(num_articles):
        art1 = Scraper('https://www.pcworld.com/search?query=nvidia&s=d&start=' +
                       str(2150-10*(i+(pages_per*iteration))) + '&submit=search&t=2', 'pcw', num_articles-j-1)

        schtuff.append(
            [art1.getArtDate(), art1.getArtTime(), art1.getArtLink(), art1.getArtFilteredText()])
        print('article ' + str((10*(i+(pages_per*iteration)))+j))
        print(len(schtuff[j][3]))
for i in range(len(schtuff)):
    newSheet.append(schtuff[i])

newWB.save(file)

print("Finished iteration: " + str(iteration))
print('Elapsed Time: ' + str(time.time()-start_time))
