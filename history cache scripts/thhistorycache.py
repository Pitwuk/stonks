import requests
from bs4 import BeautifulSoup
import openpyxl as xl
import time

from Scraper import Scraper

# This program saves the date, time, and text contents of all news articles with the search result for
# nvidia on tech radar into an xlsx file

start_time = time.time()
schtuff = []

# 0-335
num_pages = 335
iteration = 4
pages_per = 67
file = 'NVDA\\tomshardwareNVDAmasterhistory.xlsx'
worksheet = xl.load_workbook(file).active
newWB = xl.Workbook()
newSheet = newWB.active
for row in worksheet.iter_rows():
    row_arr = []
    for i in range(4):
        row_arr.append(row[i].value)
    newSheet.append(row_arr)

for i in range(pages_per):
    for j in range(20):
        art1 = Scraper('https://www.tomshardware.com/more/filter/search/nvidia/all/publishedDate/' +
                       str(num_pages-i-(pages_per*iteration))+'?searchTerm=nvidia', 'th', str(20-j))

        schtuff.append(
            [art1.getArtDate(), art1.getArtTime(), art1.getArtLink(), art1.getArtFilteredText()])
        print('article ' + str((j+((i+(pages_per*iteration))*20))) +
              '/'+str((num_pages)*20))
        print(len(schtuff[j][3]))
for i in range(len(schtuff)):
    newSheet.append(schtuff[i])

newWB.save(file)

print("Finished iteration: " + str(iteration))
print('Elapsed Time: ' + str(time.time()-start_time))
