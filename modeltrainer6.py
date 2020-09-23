import tensorflow as tf
from tensorflow import keras
from tensorflow.keras.models import Sequential
import tensorflow_hub as hub
from tensorflow.keras import layers
import tensorflow_datasets as tfds
import numpy as np
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import csv
from NextDay import NextDay

# file paths
train_articles = 'NVDA\\NVDAmasterhistory.xlsx'
# open training set
worksheet1 = load_workbook(train_articles).active


# create dictionary of next stock market dates for articles(artDates) with the key as the index of the article in the article list of integers(article_arr)


def getDates(ws):
    artDates = []
    article_arr = []
    for row in ws.rows:
        if int(str(row[0].value)[:2]) is not 0:
            article_arr.append(str(row[3].value))
            artDates.append(str(NextDay(str(row[0].value)).getNext()))
    return np.array(article_arr), np.array(artDates)


# get the integer arrays of articles and dates in a dictionary
train_article_arr, trainartDates = getDates(worksheet1)

stockws = load_workbook('NVDA\\NVDA.xlsx', data_only=True).active
changes = {}
for row in stockws.rows:
    changes[str(row[0].value)] = str(row[1].value)

# create an array of changes for the day after the article was published
train_data = {}
train_change = []
for ind in range(len(train_article_arr)):
    # print(trainartDates[ind])
    train_change.append(float(float(changes[trainartDates[ind]])))

# change data to binary labels
train_labels = []

for i in range(len(train_change)):
    if train_change[i] > 0:
        train_labels.append(int(1))
    else:
        train_labels.append(int(0))

train_labels = np.array(train_labels)

train_data = tf.data.Dataset.from_tensor_slices(
    (train_article_arr, train_labels))
train_data = train_data.shuffle(
    len(train_labels), reshuffle_each_iteration=False)
split = int(len(train_labels)*3/4)
val_data = train_data.skip(split)
train_data = train_data.take(split)
train_data = train_data.shuffle(split, reshuffle_each_iteration=True)
val_data = val_data.shuffle(len(train_labels)-split,
                            reshuffle_each_iteration=True)

train_examples_batch, train_labels_batch = next(iter(train_data.batch(10)))
print(train_labels_batch)

embedding = "https://tfhub.dev/google/tf2-preview/gnews-swivel-20dim/1"
hub_layer = hub.KerasLayer(embedding, input_shape=[],
                           dtype=tf.string, trainable=True)

model = Sequential([
    hub_layer,
    layers.Conv1D(128, 5, activation='relu'),
    layers.MaxPooling1D(5),
    layers.Conv1D(128, 5, activation='relu'),
    layers.MaxPooling1D(5),
    layers.Conv1D(128, 5, activation='relu'),
    layers.MaxPooling1D(35),
    layers.Flatten(),
    layers.Dense(128, activation='relu'),
    layers.Dense(1)
])
model.summary()

model.compile(optimizer='adam',
              loss=tf.keras.losses.BinaryCrossentropy(from_logits=True),
              metrics=['accuracy'])

mod_weights = model.get_weights()

greatest = 0.0
for i in range(50):
    EPOCHS = 1000
    early_stop = keras.callbacks.EarlyStopping(monitor='val_loss', patience=10)

    history = model.fit(train_data.shuffle(10000).batch(20), epochs=EPOCHS, validation_data=val_data.batch(20),
                        verbose=1, callbacks=[early_stop, tfdocs.modeling.EpochDots()])

    results = model.evaluate(val_data.batch(20), verbose=2)
    if results[1] > greatest:
        greatest = results[1]
    model.set_weights(mod_weights)
    print('Trial: ' + str(i))
print(greatest)
