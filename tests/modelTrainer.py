import tensorflow as tf
from tensorflow import keras
from tensorflow.keras import layers
import tensorflow_docs as tfdocs
import tensorflow_docs.plots
import tensorflow_docs.modeling
import numpy as np
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import csv
from NextDay import NextDay

# file paths
train_articles = 'NVDAmasterhistory.xlsx'
test_articles = 'nvidiaArticles7.xlsx'
dic = 'NVDAmasterdict.csv'

# set the number of words per dictionary entry
context_len = 3

# load dictionary into memory
reader = csv.reader(open(dic, 'r', encoding='utf-8'))
runtime_dic = {}
count = 4

runtime_dic['<PAD>'] = 0
runtime_dic['<UNK>'] = 1

for rows in reader:
    if int(rows[0]) > 0:
        runtime_dic[str(rows[1])] = count
        count += 1
    else:
        pass

decode_dic = {y: x for x, y in runtime_dic.items()}
# convert articles to integer lists and return the list


def toNums(article_text):
    arr = str(article_text).split(' ')
    num_arr = []
    for i in range(len(arr)-context_len):
        strang = ' '
        if strang.join(arr[i:i+context_len]) in runtime_dic:
            num_arr.append(runtime_dic[strang.join(arr[i:i+context_len])])
        else:
            num_arr.append(1)
    return num_arr


# open training set
worksheet1 = load_workbook(train_articles).active
# open the test set
worksheet2 = load_workbook(test_articles).active

# create dictionary of next stock market dates for articles(artDates) with the key as the index of the article in the article list of integers(article_arr)


def getDates(ws):
    artDates = {}
    article_arr = []
    for row in ws.rows:
        # print(row)
        if int(str(row[0].value)[:2]) is not 0:
            # print(str(NextDay(str(row[0].value)).getNext()))
            article_arr.append(toNums(str(row[3].value)))
            artDates[int(len(article_arr)-1)] = str(
                NextDay(str(row[0].value)).getNext())
    return article_arr, artDates,

# print(artDate[str(worksheet[1][3].value)])


# get the integer arrays of articles and dates in a dictionary
train_article_arr, trainartDates = getDates(worksheet1)
test_article_arr, testartDates = getDates(worksheet2)


# load NVDA stock history and create a dictionary of date:change
stockws = load_workbook('NVDA\\NVDA.xlsx', data_only=True).active
changes = {}
for row in stockws.rows:
    changes[str(row[0].value)[:10]] = str(row[1].value)

# create an array of changes for the day after the article was published
train_data = {}
train_change = []
for ind in range(len(train_article_arr)):
    train_change.append(float("{0:.6f}".format(
        float(changes[trainartDates[ind]]))))
# create test array of changes
test_data = {}
test_change = []
for ind in range(len(test_article_arr)):
    test_change.append(float("{0:.6f}".format(
        float(changes[testartDates[ind]]))))


# plot histogram of changes
train_frame = pd.DataFrame(train_change)
# sns.pairplot(train_frame, diag_kind="kde")
# plt.show()

# 1 var stats of changes
train_stats = train_frame.describe()
train_stats = train_stats.transpose()
print(train_stats)


def decode(encodedlist):
    decode_str = ''
    for i in range(len(encodedlist)):
        decode_str += decode_dic[encodedlist[i]] + ' '
    return decode_str


# print('Article: ' + str(decode(train_article_arr[0])))
# print('Change: ' + str(train_change[0]))

# convert the lists into a tensor dataset
def labeler(example, index):
    return example, tf.cast(index, tf.int64)


normed_train_change = []
for i in range(len(train_change)):
    nchange = 0
    if train_change[1] > 0:
        nchange = 1
    normed_train_change.append(int(nchange))

normed_test_change = []
for i in range(len(test_change)):
    nchange = 0
    if test_change[1] > 0:
        nchange = 1
    normed_test_change.append(int(nchange))

# x = tf.ragged.constant(train_article_arr)
# y = tf.constant(normed_train_change)
# w = tf.ragged.constant(test_article_arr)
# z = tf.constant(normed_test_change)
train_article_arr = keras.preprocessing.sequence.pad_sequences(
    train_article_arr, value=runtime_dic["<PAD>"], padding="post", maxlen=200)
test_article_arr = keras.preprocessing.sequence.pad_sequences(
    test_article_arr, value=runtime_dic["<PAD>"], padding="post", maxlen=200)
# x = tf.convert_to_tensor(train_article_arr)
# w = tf.convert_to_tensor(test_article_arr)
x = np.array(train_article_arr[:2000])
y = np.array(normed_train_change[:2000])
w = np.array(train_article_arr[2000:])
z = np.array(normed_train_change[2000:])
a = np.array(test_article_arr)
b = np.array(normed_test_change)

train_dataset = tf.data.Dataset.from_tensor_slices(
    (x, y))
test_dataset = tf.data.Dataset.from_tensor_slices(
    (w, z))
test_dataset_2 = tf.data.Dataset.from_tensor_slices(
    (a, b))
print(x.shape)
for train_example, train_label in train_dataset.take(1):
    print('Encoded text:', train_example.shape)
    print('Label:', train_label.shape)

# BATCH_SIZE = 64
# SHUFFLE_BUFFER_SIZE = 100

# train_dataset = train_dataset.shuffle(SHUFFLE_BUFFER_SIZE).batch(BATCH_SIZE)
# test_dataset = test_dataset.batch(BATCH_SIZE)


model = tf.keras.Sequential([
    tf.keras.layers.Flatten(input_shape=((200, ), ())),
])
for units in [64, 64]:
    model.add(tf.keras.layers.Dense(units, activation='relu'))
model.add(tf.keras.layers.Dense(1))

model.compile(optimizer='adam',
              loss='binary_crossentropy',
              metrics=['accuracy'])


model.fit(train_dataset, epochs=3, validation_data=test_dataset)
results = model.evaluate(test_dataset)
print(results)

model.save("testmodel.h5")

# for test_example, test_label in test_dataset_2.take(3):
#     print(test_example.shape)
#     predict = model.predict(test_example)
#     print("Review: ")
#     print(decode(test_review))
#     print("Prediction: " + str(predict))
#     print("Actual: " + str(test_label))
