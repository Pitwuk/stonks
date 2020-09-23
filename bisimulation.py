import openpyxl as xl
import numpy as np
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
import math
import scipy.stats as stats

from BinaryModelTrainer import ModelTrainer
from NextDay import NextDay


# open master article history
mast = xl.load_workbook('NVDA\\NVDAmasterhistory.xlsx').active

# load all stock changes
stockws = xl.load_workbook('NVDA\\NVDA.xlsx', data_only=True).active
changes = {}
for row in stockws.rows:
    changes[str(row[0].value)] = str(row[1].value)

# split the data at start date of simulation (2010)
past_dates = []
past_arts = []
future_dates = []
future_arts = []
count = 0
for row in mast.iter_rows():
    if int(row[0].value[6:10]) < 2019:
        past_dates.append(str(row[0].value))
        past_arts.append(str(row[3].value))
    else:
        if count < 10:
            count += 1
            future_dates.append(str(row[0].value))
            future_arts.append(str(row[3].value))

# train the model with the data before start date
model = ModelTrainer(past_dates, past_arts)


# check if the market is open the following day
def marketOpen(today):
    if NextDay(today).getNext() == NextDay(today).getNextCalDay():
        return True
    return False

# add the current days from all days until the stock market is open the following day


def addCurrent():
    global count
    today = future_dates[0]
    while len(future_dates) > 0 and future_dates[0] == today:
        current_dates.append(future_dates.pop(0))
        current_arts.append(future_arts.pop(0))
    if not marketOpen(today) and len(future_dates) > 0:
        addCurrent()

# get the actual change of the next day


def getChange():
    global changes
    return float(changes[str(NextDay(current_dates[-1]).getNext())])


# add current dates - predict change and store it - store actual change - retrain - loop
current_dates = []
current_arts = []
correct = 0
total = 0

while len(future_dates) > 0:
    addCurrent()
    predictions = model.predict(current_arts)
    print(predictions)
    mean = np.mean(predictions)
    print(mean)
    if mean > 0:
        predictions = 1
    else:
        predictions = 0
    print('Actual: ' + str(getChange()))
    if getChange() >= 0.5:
        actual = 1
    else:
        actual = 0
    if predictions == actual:
        correct += 1
    total += 1
    print(current_dates[0])

    # retrain the model with the new data
    model.retrain(current_dates, current_arts)

    for i in range(len(current_dates)):
        past_dates.append(current_dates.pop(0))
        past_arts.append(current_arts.pop(0))


print("% accuracy " + str(correct/total))
print(correct)
print(total)
