import openpyxl as xl

file = 'NVDA\\tomshardwareNVDAmasterhistory.xlsx'
worksheet = xl.load_workbook(file).active
newWB = xl.Workbook()
newSheet = newWB.active

# create dictionary of next stock market dates for articles(artDates) with the key of the index of the article in the article list of integers(article_arr)
count = 0
for row in worksheet.iter_rows():
    row_arr = []
    # print(row)
    if len(str(row[3].value)) > 10:
        count += 1
        for i in range(4):
            row_arr.append(row[i].value)
        newSheet.append(row_arr)
print(count)
# print(artDate[str(worksheet[1][3].value)])
newWB.save(file)
