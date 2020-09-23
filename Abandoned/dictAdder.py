from Dictionary import Dictionary
from openpyxl import load_workbook

# load xlsx file with full article text in it with openpyxl
# worksheet = load_workbook('NVDA\\nvidiaArticlesHistoryTR.xlsx').active
worksheet = load_workbook('nvidiaArticles7.xlsx').active

for row in worksheet.rows:
    # initialize the dictionary with a single article to add to to dictionary
    test = Dictionary('NVDAmasterdict.csv', str(row[3].value))
    print(row)
#     # add the initialized article to the dictionary in groupings of [param] words
    test.addWordsSort(3)
# print the total dictionary
test.printDict()
# dic = load_workbook('NVDAmasterdict.xlsx').active
# for row in dic.rows:
#     for cell in row:
#         print(cell.value)
# for row in worksheet.rows:
#     str(row[0].value+' '+row[1].value) = Dictionary('NVDAmasterdict.xlsx', str(row[3].value))
