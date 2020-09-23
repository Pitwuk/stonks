import tensorflow as tf
from tensorflow import keras
from tensorflow.keras.models import Sequential
from tensorflow.keras import layers
from tensorflow.keras.layers import Dense, Dropout, LSTM
import tensorflow_hub as hub
import numpy as np
from openpyxl import load_workbook
from NextDay import NextDay

# file paths
train_articles = 'NVDA\\NVDAmasterhistory.xlsx'
# open training set
worksheet1 = load_workbook(train_articles).active


# create dictionary of next stock market dates for articles(artDates) with the key as the index of the article in the article list of integers(article_arr)


def getDates(ws):
    artDates = []
    article_arr = []
    for row in ws.rows:
        if int(str(row[0].value)[:2]) is not 0:
            article_arr.append(str(row[3].value))
            artDates.append(str(NextDay(str(row[0].value)).getNext()))
    return np.array(article_arr), np.array(artDates)


# get the integer arrays of articles and dates in a dictionary
train_article_arr, trainartDates = getDates(worksheet1)

stockws = load_workbook('NVDA\\NVDA.xlsx', data_only=True).active
changes = {}
for row in stockws.rows:
    changes[str(row[0].value)] = str(row[1].value)

# create an array of changes for the day after the article was published
train_data = {}
train_change = []
for ind in range(len(train_article_arr)):
    print(trainartDates[ind])
    train_change.append(float(float(changes[trainartDates[ind]])))

# change data to binary labels
train_labels = []

for i in range(len(train_change)):
    if train_change[i] > 0:
        train_labels.append(int(1))
    else:
        train_labels.append(int(0))

train_labels = np.array(train_labels)

embed = hub.load(
    "https://tfhub.dev/google/nnlm-en-dim128-with-normalization/2")
embeddings = embed(train_article_arr)
# print(embeddings[0])
embeddings = tf.reshape(embeddings, [embeddings.shape[0], 128, 1])
print(embeddings.shape[1:])
# print(embeddings[0])

train_data = tf.data.Dataset.from_tensor_slices(
    (embeddings, train_labels))
train_data = train_data.shuffle(
    len(train_labels), reshuffle_each_iteration=False)
split = int(len(train_labels)*3/4)
val_data = train_data.skip(split)
train_data = train_data.take(split)
train_data = train_data.shuffle(split, reshuffle_each_iteration=True)
val_data = val_data.shuffle(len(train_labels)-split,
                            reshuffle_each_iteration=True)


model = Sequential([
    layers.Conv1D(128, 5, input_shape=(
        embeddings.shape[1:]), activation='relu'),
    layers.MaxPooling1D(5),
    layers.Conv1D(128, 5, activation='relu'),
    layers.MaxPooling1D(5),
    layers.Conv1D(128, 5, activation='relu'),
    layers.MaxPooling1D(35),
    layers.Flatten(),
    Dense(128, activation='relu'),
    Dense(1)
])

model.compile(loss='binary_crossentropy',
              optimizer='adam', metrics=['accuracy'])

EPOCHS = 20

history = model.fit(train_data, epochs=EPOCHS,
                    validation_data=val_data, verbose=1, batch_size=32)
