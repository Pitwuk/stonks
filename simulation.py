import openpyxl as xl
import numpy as np
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
import math
import scipy.stats as stats

from ModelTrainer import ModelTrainer
from NextDay import NextDay
from Predictor import Predictor

# open master article history
mast = xl.load_workbook('NVDA\\NVDAmasterhistory.xlsx').active

# load all stock changes
stockws = xl.load_workbook('NVDA\\NVDA.xlsx', data_only=True).active
changes = {}
for row in stockws.rows:
    changes[str(row[0].value)] = str(row[1].value)

# split the data at start date of simulation (2010)
past_dates = []
past_arts = []
future_dates = []
future_arts = []
count = 0
for row in mast.iter_rows():
    if int(row[0].value[6:10]) < 2019:
        past_dates.append(str(row[0].value))
        past_arts.append(str(row[3].value))
    else:
        if count < 10:
            count += 1
            future_dates.append(str(row[0].value))
            future_arts.append(str(row[3].value))

# train the model with the data before start date
mae = []
model = ModelTrainer(past_dates, past_arts)
mae.append(model.getMAE())


# check if the market is open the following day
def marketOpen(today):
    if NextDay(today).getNext() == NextDay(today).getNextCalDay():
        return True
    return False

# add the current days from all days until the stock market is open the following day


def addCurrent():
    global count
    today = future_dates[0]
    while len(future_dates) > 0 and future_dates[0] == today:
        current_dates.append(future_dates.pop(0))
        current_arts.append(future_arts.pop(0))
    if not marketOpen(today) and len(future_dates) > 0:
        addCurrent()

# get the actual change of the next day


def getChange():
    global changes
    return float(changes[str(NextDay(current_dates[-1]).getNext())])


# add current dates - predict change and store it - store actual change - retrain - loop
current_dates = []
current_arts = []
predicted_change = []
actual_change = []
mar_err = []
pred = Predictor()
while len(future_dates) > 0:
    addCurrent()
    predictions = pred.predict(current_arts)
    std_dev = np.std(predictions)
    mean = np.mean(predictions)
    predicted_change.append(mean)
    actual_change.append(getChange())
    print(current_dates[0])

    model.retrain(current_dates, current_arts)
    # retrain the model with the new data
    mae.append(model.getMAE())
    # calculate the 1 sample t interval for the prediction with a 90% confidence level
    if len(current_dates) > 1:
        mar_err.append(float((abs(stats.t.ppf(0.05, df=(len(current_dates)-1)))
                              * (std_dev/math.sqrt(len(current_dates))))))
    else:
        mar_err.append(0)
    print('Margin of Error: ' + str(mar_err[-1]))
    for i in range(len(current_dates)):
        past_dates.append(current_dates.pop(0))
        past_arts.append(current_arts.pop(0))

error = []
for i in range(len(predicted_change)):
    error.append(abs(predicted_change[i]-actual_change[i]))

big_error = []
count = 0
for i in range(len(error)):
    if float(error[i]) > float(mar_err[i]):
        big_error.append(1)
        count += 1
    else:
        big_error.append(0)
print("% of predictions outside of the interval: " + str(count/len(big_error)))

data = pd.DataFrame(list(zip(predicted_change, actual_change, error, mae)),
                    columns=['predicted', 'actual', 'error', 'mae'])
data['trial'] = data.index.tolist()
sns.lmplot('trial', 'mae', data=data, ci=None)
plt.show()
sns.lmplot('predicted', 'actual', data=data)
plt.show()

changeWB = xl.Workbook()
changesheet = changeWB.active
for row in data.values.tolist():
    changesheet.append(row)
changeWB.save('changes.xlsx')
