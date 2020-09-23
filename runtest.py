import time
from datetime import date
from datetime import timedelta
import math
import keyboard
from alpha_vantage.timeseries import TimeSeries
import pandas as pd
import openpyxl as xl
from Scraper import Scraper
from NextDay import NextDay

# alpha-vantage realtime stock prices
key = 'P6O4FHCBVPESJ74D'
ts = TimeSeries(key)


site_dic = {'tr': 'https://www.techradar.com/more/filter/search/nvidia/news/publishedDate/1?searchTerm=nvidia',
            'pcw': 'https://www.pcworld.com/search?query=nvidia&s=d',
            'vg': 'https://www.theverge.com/search?order=date&q=nvidia&type=Article',
            'th': 'https://www.tomshardware.com/filter/search?searchTerm=nvidia&sortBy=publishedDate&articleType=all',
            'pcg': 'https://www.pcgamer.com/more/filter/search/nvidia/all/publishedDate/1/?searchTerm=nvidia',
            'aat': 'https://www.anandtech.com/SearchResults?q=nvidia'}


# checks the given website for the [num] most recent articles and adds them to the xlsx file if they were published today and aren't already in there
def getNew(site, day, num):
    file = 'NVDA\\currentNVDA.xlsx'
    worksheet = xl.load_workbook(file).active
    newWB = xl.Workbook()
    newSheet = newWB.active
    link_storage = []
    for row in worksheet.iter_rows():
        row_arr = []
        for i in range(4):
            row_arr.append(str(row[i].value))
        link_storage.append(row_arr[2])
        newSheet.append(row_arr)
    count = 0
    for i in range(num):
        art1 = Scraper(site_dic[site], site, i+1)
        if art1.getArtDate() == day and art1.getArtLink() not in link_storage:
            print('Found ' + site + ' Article')
            count += 1
            newSheet.append([art1.getArtDate(), art1.getArtTime(
            ), art1.getArtLink(), art1.getArtFilteredText()])
    newWB.save(file)
    print('Finished ' + site + ' Scrape: ' + str(count) + ' Articles')


def dumpYesterday(day):
    mast = 'NVDA\\NVDAmasterhistory.xlsx'
    cur = 'NVDA\\currentNVDA.xlsx'
    history = xl.load_workbook(mast).active
    current = xl.load_workbook(cur).active
    newWB = xl.Workbook()
    newSheet = newWB.active
    currWB = xl.Workbook()
    currSheet = currWB.active
    art_storage = []
    for row in history.iter_rows():
        row_arr = []
        for i in range(4):
            row_arr.append(row[i].value)
        newSheet.append(row_arr)
    count = 0
    for row in current.iter_rows():
        row_arr = []
        curr_arr = []
        if not str(row[0].value) == day and len(str(row[3].value)) > 10:
            for i in range(4):
                row_arr.append(row[i].value)
            newSheet.append(row_arr)
            count += 1
        elif len(str(row[3].value)) > 10:
            for i in range(4):
                curr_arr.append(row[i].value)
            currSheet.append(curr_arr)
    newWB.save(mast)
    currWB.save(cur)
    print(str(count) + " Articles Dumped Successfully")

# add today's nvda stock change


def addStonk(day):
    # get change
    nvda, meta = ts.get_daily(symbol='NVDA')
    date = day[6:] + '-' + day[:2] + '-' + day[3:5]
    day_stats = nvda[date]
    change = ((float(day_stats['4. close']) -
               float(day_stats['1. open']))/float(day_stats['1. open']))*100
    print('Open: ' + str(day_stats['1. open']))
    print('Close: ' + str(day_stats['4. close']))
    print('Change: ' + str(change))

    # write change
    file = 'NVDA\\NVDA.xlsx'
    worksheet = xl.load_workbook(file, data_only=True).active
    newWB = xl.Workbook()
    newSheet = newWB.active
    for row in worksheet.iter_rows():
        row_arr = []
        dt = str(row[0].value)
        row_arr.append(dt)
        row_arr.append(str(row[1].value))
        newSheet.append(row_arr)
    if not dt == day:
        new_row = [day, str(change)]
        newSheet.append(new_row)
        newWB.save(file)
        print("Successfully Added NVDA Change")
    else:
        print('Already added')

# checks if the stock market is open on the given day


def marketOpen(yesterday, today):
    print(NextDay(yesterday).getNext())
    if NextDay(yesterday).getNext() == today:
        return True
    print('Stock Market Closed Today')
    return False


num_yest = 0
num_curr = 0

yesterday = '04/07/2020'
day = '04/08/2020'

print(day)

print('\n')
try:
    getNew('tr', yesterday, num_yest)
except:
    print('Error getting Tech Radar Yesterday Articles')
try:
    getNew('pcw', yesterday, num_yest)
except:
    print('Error getting PC World Yesterday Articles')
try:
    getNew('vg', yesterday, num_yest)
except:
    print('Error getting The Verge Yesterday Articles')
try:
    getNew('th', yesterday, num_yest)
except:
    print('Error getting Tom\'s Hardware Yesterday Articles')
try:
    getNew('pcg', yesterday, num_yest)
except:
    print('Error getting PC Gamer Yesterday Articles')
try:
    getNew('aat', yesterday, num_yest)
except:
    print('Error getting AnandTech Yesterday Articles')
yest_check = True
print("Got All Yesterday Articles")

# get most recent articles
print('\n')

try:
    getNew('tr', day, num_curr)
except:
    print('Error getting Tech Radar Articles')
try:
    getNew('pcw', day, num_curr)
except:
    print('Error getting PC World Articles')
try:
    getNew('vg', day, num_curr)
except:
    print('Error getting The Verge Articles')
try:
    getNew('th', day, num_curr)
except:
    print('Error getting Tom\'s Hardware Articles')
try:
    getNew('pcg', day, num_curr)
except:
    print('Error getting PC Gamer Articles')
try:
    getNew('aat', day, num_curr)
except:
    print('Error getting AnandTech Articles')
if marketOpen(yesterday, day):
    dumpYesterday(day)
    addStonk(day)
