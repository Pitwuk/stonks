import tensorflow as tf
from tensorflow import keras
from tensorflow.keras.models import Sequential
from tensorflow.keras import layers
from tensorflow.keras.layers import Dense, Dropout, LSTM, Embedding
from tensorflow.keras.preprocessing.text import Tokenizer
from tensorflow.keras.preprocessing.sequence import pad_sequences
import numpy as np
from openpyxl import load_workbook
from NextDay import NextDay

# file paths
train_articles = 'NVDA\\NVDAmasterhistory.xlsx'
# open training set
worksheet1 = load_workbook(train_articles).active


# create dictionary of next stock market dates for articles(artDates) with the key as the index of the article in the article list of integers(article_arr)


def getDates(ws):
    artDates = []
    article_arr = []
    for row in ws.rows:
        if int(str(row[0].value)[:2]) is not 0:
            article_arr.append(str(row[3].value))
            artDates.append(str(NextDay(str(row[0].value)).getNext()))
    return np.array(article_arr), np.array(artDates)


# get the integer arrays of articles and dates in a dictionary
train_article_arr, trainartDates = getDates(worksheet1)

stockws = load_workbook('NVDA\\NVDA.xlsx', data_only=True).active
changes = {}
for row in stockws.rows:
    changes[str(row[0].value)] = str(row[1].value)

# create an array of changes for the day after the article was published
train_data = {}
train_change = []
for ind in range(len(train_article_arr)):
    # print(trainartDates[ind])
    train_change.append(float(float(changes[trainartDates[ind]])))

# change data to binary labels
train_labels = []

for i in range(len(train_change)):
    if train_change[i] > 0:
        train_labels.append(int(1))
    else:
        train_labels.append(int(0))

train_labels = np.array(train_labels)

train_labels = train_labels[:3000]
train_article_arr = train_article_arr[:3000]
# tokenize the articles
tokenizer = Tokenizer(num_words=20000, oov_token="<OOV>")
tokenizer.fit_on_texts(train_article_arr)
word_index = tokenizer.word_index
print(len(word_index))
sequences = tokenizer.texts_to_sequences(train_article_arr)
padded = pad_sequences(sequences)
print(len(padded[0]))

# # create and split datasets and shuffle
# train_data = tf.data.Dataset.from_tensor_slices(
#     (padded, train_labels))
# train_data = train_data.shuffle(
#     len(train_labels), reshuffle_each_iteration=False)
# split = int(len(train_labels)*3/4)
# val_data = train_data.skip(split)
# train_data = train_data.take(split)
# train_data = train_data.shuffle(split, reshuffle_each_iteration=True)
# val_data = val_data.shuffle(len(train_labels)-split,
#                             reshuffle_each_iteration=True)
# test_data = padded[int(len(padded)*3/4):int(len(padded))]
# test_labels = train_labels[int(len(padded)*3/4):int(len(padded))]
# train_data = padded[:int(len(padded)*3/4)]
# train_labels = train_labels[:int(len(padded)*3/4)]
# print(padded.shape)
test_data = padded[::2]
test_labels = train_labels[::2]
train_data = padded[1::2]
train_labels = train_labels[1::2]
print(padded.shape)
# print(test_data[0])
# print(test_labels[0])

# model = Sequential([
#     Embedding(len(word_index), 16, input_length=len(padded[0])),
#     layers.GlobalAveragePooling1D(),
#     Dense(16, activation='relu'),
#     Dense(1, activation='sigmoid')
# ])

# model = Sequential([
#     Embedding(len(word_index), 32, input_length=len(padded[0])),
#     LSTM(32, return_sequences=True),
#     # Dropout(0.2),
#     # LSTM(32),
#     # Dropout(0.1),
#     Dense(32, activation='relu'),
#     # Dropout(0.2),
#     Dense(1, activation='sigmoid')
# ])
model = Sequential([
    Embedding(len(word_index)+1, 128, input_length=len(padded[0])),
    layers.Conv1D(128, 5, activation='relu'),
    layers.MaxPooling1D(5),
    layers.Conv1D(128, 5, activation='relu'),
    layers.MaxPooling1D(5),
    layers.Conv1D(128, 5, activation='relu'),
    layers.MaxPooling1D(35),
    layers.Flatten(),
    Dense(128, activation='relu'),
    Dense(1)
])

model.compile(loss='binary_crossentropy',
              optimizer='adam', metrics=['accuracy'])

EPOCHS = 100

early_stop = keras.callbacks.EarlyStopping(monitor='val_loss', patience=10)

history = model.fit(train_data, train_labels, epochs=EPOCHS,
                    validation_data=(test_data, test_labels), verbose=1, batch_size=16, callbacks=[early_stop])
