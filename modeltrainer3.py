import tensorflow as tf
from tensorflow import keras
from keras import layers
import tensorflow_docs as tfdocs
import tensorflow_docs.plots
import tensorflow_docs.modeling
import tensorflow_hub as hub
import tensorflow_datasets as tfds
import numpy as np
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import csv
from NextDay import NextDay

# file paths
train_articles = 'NVDA\\NVDAmasterhistory.xlsx'
# open training set
worksheet1 = load_workbook(train_articles).active


# create dictionary of next stock market dates for articles(artDates) with the key as the index of the article in the article list of integers(article_arr)


def getDates(ws):
    artDates = []
    article_arr = []
    for row in ws.rows:
        if int(str(row[0].value)[:2]) is not 0:
            article_arr.append(str(row[3].value))
            artDates.append(str(NextDay(str(row[0].value)).getNext()))
    return np.array(article_arr), np.array(artDates)


# get the integer arrays of articles and dates in a dictionary
train_article_arr, trainartDates = getDates(worksheet1)

stockws = load_workbook('NVDA\\NVDA.xlsx', data_only=True).active
changes = {}
for row in stockws.rows:
    changes[str(row[0].value)] = str(row[1].value)

# create an array of changes for the day after the article was published
train_data = {}
train_change = []
for ind in range(len(train_article_arr)):
    print(trainartDates[ind])
    train_change.append(float(float(changes[trainartDates[ind]])))

# change data to binary labels
train_labels = []

for i in range(len(train_change)):
    if train_change[i] > 0:
        train_labels.append(int(1))
    else:
        train_labels.append(int(0))

train_labels = np.array(train_change)

train_data = tf.data.Dataset.from_tensor_slices(
    (train_article_arr, train_labels))
train_data = train_data.shuffle(
    len(train_labels), reshuffle_each_iteration=False)
split = int(len(train_labels)*3/4)
val_data = train_data.skip(split)
train_data = train_data.take(split)
train_data = train_data.shuffle(split, reshuffle_each_iteration=True)
val_data = val_data.shuffle(len(train_labels)-split,
                            reshuffle_each_iteration=True)

# embedding = "https://tfhub.dev/google/tf2-preview/gnews-swivel-20dim/1"
# hub_layer = hub.KerasLayer(embedding, output_shape=[20], input_shape=[],
#                            dtype=tf.string, trainable=True)

model = tf.keras.Sequential()
model.add(hub.KerasLayer(
    "https://tfhub.dev/google/nnlm-en-dim50-with-normalization/2", input_shape=[], output_shape=[50], dtype=tf.string, trainable=True))
# model.add(tf.keras.layers.Bidirectional(
#     tf.keras.layers.LSTM(64, input_shape=[None, 128], return_sequences=True)))
# model.add(tf.keras.layers.Dense(32, activation='relu'))
model.add(tf.keras.layers.LSTM(50, input_shape=(None, 50)))
model.add(tf.keras.layers.Dense(32, activation='relu'))
model.add(tf.keras.layers.Dense(1, activation='sigmoid'))

model.summary()

model.compile(loss='mse', optimizer=tf.keras.optimizers.RMSprop(0.001),
              metrics=['mae', 'mse'])

mod_weights = model.get_weights()

lowest = 999.99
for i in range(10):
    EPOCHS = 1000
    early_stop = keras.callbacks.EarlyStopping(monitor='val_mae', patience=2)

    history = model.fit(train_data.batch(20), epochs=EPOCHS, validation_data=val_data.batch(20),
                        verbose=2, callbacks=[early_stop, tfdocs.modeling.EpochDots()])

    loss, mae, mse = model.evaluate(val_data.batch(20), verbose=2)
    print("Testing set Mean Abs Error: {:5.2f}%".format(mae))
    if float(mae) < lowest:
        lowest = float(mae)
        # model.save("model.h5")
    model.set_weights(mod_weights)
    print('Trial: ' + str(i))
print(lowest)